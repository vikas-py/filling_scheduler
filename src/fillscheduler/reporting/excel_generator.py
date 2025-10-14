"""
Excel Report Generator

Generates formatted Excel workbooks from schedule data using openpyxl.
"""

from typing import Dict, Any, Optional
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(
    schedule_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate a formatted Excel workbook from schedule data.
    
    Args:
        schedule_data: Dictionary containing schedule information
        output_path: Optional path to save Excel file
    
    Returns:
        Excel file content as bytes
    """
    schedule = schedule_data['schedule']
    results = schedule_data['results']
    activities = schedule_data.get('activities', [])
    config = schedule_data.get('config', {})
    
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = 'Schedule Report'
    ws_summary['A1'].font = Font(size=16, bold=True, color='1976D2')
    ws_summary['A2'] = schedule.get('name', 'Unnamed Schedule')
    ws_summary['A2'].font = Font(size=12, italic=True)
    
    # KPIs
    row = 4
    ws_summary[f'A{row}'] = 'Key Metrics'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    kpis = [
        ('Makespan', f"{results.get('makespan', 0):.2f} hours"),
        ('Utilization', f"{results.get('utilization', 0) * 100:.1f}%"),
        ('Changeovers', results.get('changeovers', 0)),
        ('Lots Scheduled', results.get('lots_scheduled', 0)),
        ('Strategy', schedule.get('strategy', 'Unknown')),
        ('Status', schedule.get('status', 'Unknown')),
    ]
    
    for label, value in kpis:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Auto-size columns
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25
    
    # Sheet 2: Activities
    ws_activities = wb.create_sheet("Activities")
    
    # Headers
    headers = ['Lot ID', 'Filler', 'Type', 'Start (h)', 'End (h)', 'Duration (h)', 'Lot Type']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_activities.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row_idx, activity in enumerate(activities, 2):
        ws_activities.cell(row=row_idx, column=1, value=activity.get('lot_id', ''))
        ws_activities.cell(row=row_idx, column=2, value=activity.get('filler_id', ''))
        ws_activities.cell(row=row_idx, column=3, value=activity.get('kind', 'FILL'))
        ws_activities.cell(row=row_idx, column=4, value=round(activity.get('start_time', 0), 2))
        ws_activities.cell(row=row_idx, column=5, value=round(activity.get('end_time', 0), 2))
        ws_activities.cell(row=row_idx, column=6, value=round(activity.get('duration', 0), 2))
        ws_activities.cell(row=row_idx, column=7, value=activity.get('lot_type', ''))
        
        # Apply borders
        for col_idx in range(1, 8):
            ws_activities.cell(row=row_idx, column=col_idx).border = border
    
    # Auto-size columns
    for col_idx in range(1, 8):
        ws_activities.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Sheet 3: Configuration
    ws_config = wb.create_sheet("Configuration")
    
    ws_config['A1'] = 'Configuration Parameters'
    ws_config['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for key, value in config.items():
        ws_config[f'A{row}'] = key.replace('_', ' ').title()
        ws_config[f'A{row}'].font = Font(bold=True)
        ws_config[f'B{row}'] = str(value)
        row += 1
    
    ws_config.column_dimensions['A'].width = 30
    ws_config.column_dimensions['B'].width = 20
    
    # Save or return bytes
    if output_path:
        wb.save(output_path)
        with open(output_path, 'rb') as f:
            return f.read()
    else:
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.read()


if __name__ == '__main__':
    # Test function
    import random
    
    activities = []
    for i in range(20):
        filler_id = random.randint(1, 3)
        start = random.uniform(0, 40)
        duration = random.uniform(2, 8)
        activities.append({
            'lot_id': f'LOT_{i:03d}',
            'filler_id': filler_id,
            'start_time': start,
            'end_time': start + duration,
            'duration': duration,
            'kind': random.choice(['FILL', 'CLEAN', 'CHANGEOVER']),
            'lot_type': random.choice(['TypeA', 'TypeB']),
        })
    
    schedule_data = {
        'schedule': {
            'name': 'Test Schedule',
            'status': 'completed',
            'strategy': 'MILP',
        },
        'results': {
            'makespan': 48.5,
            'utilization': 0.85,
            'changeovers': 12,
            'lots_scheduled': 20,
        },
        'activities': activities,
        'config': {
            'num_fillers': 3,
            'clean_time_same': 1.5,
            'clean_time_different': 3.0,
        },
    }
    
    try:
        print("Testing Excel generation...")
        excel_bytes = generate_excel_report(schedule_data, output_path="test_report.xlsx")
        print(f"✓ Excel generated successfully! Size: {len(excel_bytes)} bytes")
        print("✓ Saved to: test_report.xlsx")
    except Exception as e:
        print(f"✗ Excel generation failed: {e}")
        import traceback
        traceback.print_exc()
